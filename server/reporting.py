from __future__ import annotations

from datetime import datetime
from io import BytesIO
import logging
import os
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.pdfmetrics import registerFontFamily
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import text
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import Session

logger = logging.getLogger(__name__)

ORDER_STATUS_LABELS = {
    "new": "Нове",
    "processing": "В обробці",
    "shipped": "Відправлено",
    "delivered": "Доставлено",
    "picked_up": "Забрано",
    "cancelled": "Скасовано",
    "refunded": "Повернено",
}


REPORT_FONT_CANDIDATES = [
    os.getenv("REPORT_FONT_PATH", "").strip(),
    # Windows fonts - prioritize Unicode-capable fonts
    str(Path("C:/Windows/Fonts/segoeui.ttf")),  # Segoe UI has excellent Unicode/Cyrillic support
    str(Path("C:/Windows/Fonts/arialuni.ttf")),  # Arial Unicode MS if available
    str(Path("C:/Windows/Fonts/arial.ttf")),
    # Local project fonts
    str(Path(__file__).resolve().parent / "fonts" / "DejaVuSans.ttf"),
    # Linux fonts
    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    "/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf",
    "/usr/share/fonts/dejavu/DejaVuSans.ttf",
]

# Required glyphs for Ukrainian labels used across the report.
PDF_REQUIRED_GLYPHS = "АБВГДЕЄЖЗИІЇЙКЛМНОПРСТУФХЦЧШЩЬЮЯабвгдеєжзиіїйклмнопрстуфхцчшщьюяҐґ"


def _safe_money(value) -> str:
    try:
        return f"{float(value or 0):,.2f}".replace(",", " ")
    except (TypeError, ValueError):
        return "0.00"


def _safe_int(value, default: int = 0) -> int:
    try:
        if value is None:
            return default
        return int(value)
    except (TypeError, ValueError):
        return default


def _status_label(status_value: str) -> str:
    return ORDER_STATUS_LABELS.get(str(status_value or "new"), str(status_value or "new"))


def _format_dt(value: Any) -> str:
    if value is None:
        return "-"
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y %H:%M")
    text_value = str(value)
    if "T" in text_value:
        return text_value.replace("T", " ")[:16]
    return text_value[:16]


def _scalar_or_default(db: Session, statement, default=0):
    try:
        return db.scalar(statement) or default
    except SQLAlchemyError as error:
        logger.warning("report scalar query failed", extra={"error": str(error)})
        return default


def _rows_or_empty(db: Session, statement):
    try:
        return db.execute(statement).mappings().all()
    except SQLAlchemyError as error:
        logger.warning("report rows query failed", extra={"error": str(error)})
        return []


def collect_admin_report_data(db: Session) -> dict[str, Any]:
    generated_at = datetime.now()

    counts = {
        "categories": _scalar_or_default(db, text("SELECT COUNT(*) FROM categories"), 0),
        "products": _scalar_or_default(db, text("SELECT COUNT(*) FROM products"), 0),
        "orders": _scalar_or_default(db, text("SELECT COUNT(*) FROM orders"), 0),
        "users": _scalar_or_default(db, text("SELECT COUNT(*) FROM users"), 0),
    }

    inventory_rows = _rows_or_empty(db, text("""
        SELECT
            i.id,
            i.product_id,
            COALESCE(i.quantity, 0) AS quantity,
            COALESCE(i.min_quantity, 0) AS min_quantity,
            COALESCE(i.min_quantity_alert, i.min_quantity, 0) AS threshold,
            COALESCE(i.max_quantity, 0) AS max_quantity,
            COALESCE(i.location, '-') AS location,
            p.name AS product_name,
            p.sku AS product_sku
        FROM inventory i
        LEFT JOIN products p ON p.id = i.product_id
        ORDER BY COALESCE(i.quantity, 0) ASC, i.id ASC
    """))
    low_stock_rows = [row for row in inventory_rows if _safe_int(row["quantity"], 0) < _safe_int(row["threshold"], 0)]
    out_of_stock_rows = [row for row in inventory_rows if _safe_int(row["quantity"], 0) <= 0]
    total_stock_units = sum(max(_safe_int(row["quantity"], 0), 0) for row in inventory_rows)

    status_rows = _rows_or_empty(db, text("""
        SELECT COALESCE(status, 'new') AS status, COUNT(*) AS count
        FROM orders
        GROUP BY COALESCE(status, 'new')
        ORDER BY count DESC
    """))
    status_counts = [
        {"status": str(row["status"]), "label": _status_label(row["status"]), "count": _safe_int(row["count"], 0)}
        for row in status_rows
    ]

    paid_revenue = _scalar_or_default(db, text("""
        SELECT COALESCE(SUM(total), 0)
        FROM orders
        WHERE status IN ('delivered', 'picked_up')
    """), 0)

    latest_orders = _rows_or_empty(db, text("""
        SELECT
            id,
            user_id,
            contact_name,
            contact_phone,
            contact_email,
            delivery_city,
            delivery_address,
            COALESCE(status, 'new') AS status,
            COALESCE(total, 0) AS total,
            created_at
        FROM orders
        ORDER BY created_at DESC
        LIMIT 12
    """))

    top_products = _rows_or_empty(db, text("""
        SELECT
            id,
            name,
            sku,
            COALESCE(price, 0) AS price,
            CASE WHEN is_active IS TRUE THEN 1 ELSE 0 END AS is_active
        FROM products
        ORDER BY COALESCE(price, 0) DESC, id DESC
        LIMIT 10
    """))

    return {
        "generated_at": generated_at,
        "counts": counts,
        "total_stock_units": total_stock_units,
        "low_stock_rows": low_stock_rows,
        "out_of_stock_rows": out_of_stock_rows,
        "status_counts": status_counts,
        "paid_revenue": paid_revenue,
        "latest_orders": latest_orders,
        "top_products": top_products,
    }


def _resolve_pdf_font() -> str:
    for candidate in REPORT_FONT_CANDIDATES:
        path = Path(candidate)
        if not candidate or not path.exists():
            continue
        font_name = f"BuildShop-{path.stem}"
        try:
            font_obj = TTFont(font_name, str(path))
        except Exception as error:
            logger.warning("PDF font candidate failed to load", extra={"font_path": str(path), "error": str(error)})
            continue

        missing_chars = [ch for ch in PDF_REQUIRED_GLYPHS if ord(ch) not in font_obj.face.charToGlyph]
        if missing_chars:
            logger.warning(
                "PDF font candidate skipped due to missing glyphs",
                extra={"font_path": str(path), "missing_count": len(missing_chars)},
            )
            continue

        if font_name not in pdfmetrics.getRegisteredFontNames():
            pdfmetrics.registerFont(font_obj)
            try:
                registerFontFamily(font_name, normal=font_name, bold=font_name, italic=font_name, boldItalic=font_name)
            except Exception:
                pass
        logger.info("PDF report font selected", extra={"font_name": font_name, "font_path": str(path)})
        return font_name
    raise RuntimeError(
        "Unicode PDF font not found. Configure REPORT_FONT_PATH or install DejaVu fonts (e.g., apt-get install fonts-dejavu-core)."
    )


def _make_paragraph_style(font_name: str, size: int = 9, bold: bool = False, alignment: int | None = None):
    return ParagraphStyle(
        name=f"Report-{font_name}-{size}-{int(bold)}-{alignment or 0}",
        fontName=font_name,
        fontSize=size,
        leading=size + 2,
        textColor=colors.HexColor("#0f172a"),
        alignment=alignment if alignment is not None else 0,
    )


def _paragraph(value: Any, style: ParagraphStyle) -> Paragraph:
    text_value = "-" if value is None else str(value)
    text_value = text_value.replace("\n", "<br/>")
    return Paragraph(text_value, style)


def _styled_table(rows: list[list[Any]], font_name: str, col_widths: list[float] | None = None) -> Table:
    body = _make_paragraph_style(font_name, 8)
    header = _make_paragraph_style(font_name, 8, bold=True, alignment=TA_CENTER)
    table_data = []
    for row_index, row in enumerate(rows):
        style = header if row_index == 0 else body
        table_data.append([_paragraph(cell, style) for cell in row])

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), font_name),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("LEADING", (0, 0), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#cbd5e1")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    return table


def build_admin_report_pdf(data: dict[str, Any]) -> bytes:
    font_name = _resolve_pdf_font()
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title="BuildShop адміністративний звіт",
        author="BuildShop",
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "BuildShopTitle",
        parent=styles["Title"],
        fontName=font_name,
        fontSize=22,
        leading=26,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#0f172a"),
    )
    subtitle_style = ParagraphStyle(
        "BuildShopSubtitle",
        parent=styles["BodyText"],
        fontName=font_name,
        fontSize=10,
        leading=13,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#475569"),
    )
    section_style = ParagraphStyle(
        "BuildShopSection",
        parent=styles["Heading2"],
        fontName=font_name,
        fontSize=14,
        leading=16,
        spaceBefore=6,
        spaceAfter=6,
        textColor=colors.HexColor("#111827"),
    )
    note_style = ParagraphStyle(
        "BuildShopNote",
        parent=styles["BodyText"],
        fontName=font_name,
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#475569"),
    )

    story = [
        Paragraph("BuildShop — Адміністративний звіт", title_style),
        Spacer(1, 4 * mm),
        Paragraph(f"Дата формування: {data['generated_at'].strftime('%d.%m.%Y %H:%M')}", subtitle_style),
        Paragraph("Звіт містить ключові показники, статуси замовлень, топ-товари та ризики по складу.", subtitle_style),
        Spacer(1, 6 * mm),
    ]

    counts = data["counts"]
    story.append(Paragraph("1. Ключові показники", section_style))
    story.append(_styled_table([
        ["Показник", "Значення"],
        ["Категорії", counts["categories"]],
        ["Товари", counts["products"]],
        ["Замовлення", counts["orders"]],
        ["Користувачі", counts["users"]],
        ["Одиниць товару на складі", data["total_stock_units"]],
        ["Низький запас", len(data["low_stock_rows"])],
        ["Немає в наявності", len(data["out_of_stock_rows"])],
        ["Виторг (доставлено/забрано), грн", _safe_money(data["paid_revenue"])],
    ], font_name, [120 * mm, 65 * mm]))

    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph("2. Статуси замовлень", section_style))
    status_rows = data["status_counts"]
    if status_rows:
        story.append(_styled_table(
            [["Статус", "Кількість"]] + [[row["label"], row["count"]] for row in status_rows],
            font_name,
            [130 * mm, 55 * mm],
        ))
    else:
        story.append(Paragraph("Немає даних по замовленнях.", note_style))

    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph("3. Останні замовлення", section_style))
    latest_orders = data["latest_orders"]
    if latest_orders:
        story.append(_styled_table(
            [["№", "Дата", "Клієнт", "Статус", "Сума, грн"]] + [
                [
                    f"#{row['id']}",
                    _format_dt(row.get("created_at")),
                    row.get("contact_name") or f"user #{row.get('user_id')}",
                    _status_label(row.get("status")),
                    _safe_money(row.get("total")),
                ]
                for row in latest_orders
            ],
            font_name,
            [20 * mm, 30 * mm, 60 * mm, 45 * mm, 30 * mm],
        ))
    else:
        story.append(Paragraph("Останні замовлення відсутні.", note_style))

    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph("4. Найдорожчі товари", section_style))
    top_products = data["top_products"]
    if top_products:
        story.append(_styled_table(
            [["Товар", "SKU", "Ціна, грн", "Активний"]] + [
                [
                    row.get("name") or "-",
                    row.get("sku") or "-",
                    _safe_money(row.get("price")),
                    "Так" if _safe_int(row.get("is_active"), 0) else "Ні",
                ]
                for row in top_products
            ],
            font_name,
            [95 * mm, 35 * mm, 35 * mm, 25 * mm],
        ))
    else:
        story.append(Paragraph("Товари не знайдено.", note_style))

    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph("5. Критичні позиції складу", section_style))
    low_stock_rows = data["low_stock_rows"]
    if low_stock_rows:
        story.append(_styled_table(
            [["Товар", "SKU", "К-сть", "Поріг", "Локація"]] + [
                [
                    row.get("product_name") or "Невідомо",
                    row.get("product_sku") or "-",
                    _safe_int(row.get("quantity"), 0),
                    _safe_int(row.get("threshold"), 0),
                    row.get("location") or "-",
                ]
                for row in low_stock_rows[:30]
            ],
            font_name,
            [75 * mm, 35 * mm, 20 * mm, 25 * mm, 50 * mm],
        ))
        if len(low_stock_rows) > 30:
            story.append(Spacer(1, 2 * mm))
            story.append(Paragraph(f"Показано перші 30 позицій із {len(low_stock_rows)}.", note_style))
    else:
        story.append(Paragraph("Критичних позицій по складу немає.", note_style))

    doc.build(story)
    return buffer.getvalue()


def _auto_widths(rows: list[list[Any]], min_width: int = 10, max_width: int = 42) -> list[float]:
    widths: list[float] = []
    columns = max((len(row) for row in rows), default=0)
    for index in range(columns):
        max_len = max((len(str(row[index])) if index < len(row) and row[index] is not None else 0) for row in rows)
        width = max(min_width, min(max_width, max_len * 1.2))
        widths.append(width)
    return widths


def _style_ws(ws, header_row: int = 1):
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row in ws.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = ws[2][0].coordinate if ws.max_row > header_row else "A2"
    ws.auto_filter.ref = ws.dimensions
    for column_cells in ws.columns:
        column_letter = column_cells[0].column_letter
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 42)


def _write_table_sheet(ws, title: str, headers: list[str], rows: list[list[Any]]):
    ws.title = title
    ws.append(headers)
    for row in rows:
        ws.append(row)
    _style_ws(ws)


def build_admin_report_xlsx(data: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Підсумок"

    ws_summary["A1"] = "BuildShop — Адміністративний звіт"
    ws_summary["A1"].font = Font(size=16, bold=True, color="0F172A")
    ws_summary["A2"] = f"Дата формування: {data['generated_at'].strftime('%d.%m.%Y %H:%M')}"
    ws_summary["A3"] = "Звіт містить ключові показники, статуси замовлень, топ-товари та ризики по складу."
    ws_summary["A5"] = "Ключові показники"
    ws_summary["A5"].font = Font(size=12, bold=True)

    summary_rows = [["Показник", "Значення"]] + [
        ["Категорії", data["counts"]["categories"]],
        ["Товари", data["counts"]["products"]],
        ["Замовлення", data["counts"]["orders"]],
        ["Користувачі", data["counts"]["users"]],
        ["Одиниць товару на складі", data["total_stock_units"]],
        ["Низький запас", len(data["low_stock_rows"])],
        ["Немає в наявності", len(data["out_of_stock_rows"])],
        ["Виторг (доставлено/забрано), грн", _safe_money(data["paid_revenue"])],
    ]
    for row in summary_rows:
        ws_summary.append(row)

    # status block below summary table
    start_row = len(summary_rows) + 3
    ws_summary.cell(row=start_row, column=1, value="Статуси замовлень").font = Font(size=12, bold=True)
    ws_summary.cell(row=start_row + 1, column=1, value="Статус")
    ws_summary.cell(row=start_row + 1, column=2, value="Кількість")
    for index, row in enumerate(data["status_counts"], start=start_row + 2):
        ws_summary.cell(row=index, column=1, value=row["label"])
        ws_summary.cell(row=index, column=2, value=row["count"])
    _style_ws(ws_summary)

    ws_orders = wb.create_sheet("Замовлення")
    order_rows = [["№", "Дата", "Клієнт", "Статус", "Сума, грн"]] + [
        [
            f"#{row['id']}",
            _format_dt(row.get("created_at")),
            row.get("contact_name") or f"user #{row.get('user_id')}",
            _status_label(row.get("status")),
            float(row.get("total") or 0),
        ]
        for row in data["latest_orders"]
    ]
    for row in order_rows:
        ws_orders.append(row)
    _style_ws(ws_orders)

    ws_products = wb.create_sheet("Товари")
    product_rows = [["Товар", "SKU", "Ціна, грн", "Активний"]] + [
        [
            row.get("name") or "-",
            row.get("sku") or "-",
            float(row.get("price") or 0),
            "Так" if _safe_int(row.get("is_active"), 0) else "Ні",
        ]
        for row in data["top_products"]
    ]
    for row in product_rows:
        ws_products.append(row)
    _style_ws(ws_products)

    ws_stock = wb.create_sheet("Склад")
    stock_rows = [["Товар", "SKU", "К-сть", "Поріг", "Локація"]] + [
        [
            row.get("product_name") or "Невідомо",
            row.get("product_sku") or "-",
            _safe_int(row.get("quantity"), 0),
            _safe_int(row.get("threshold"), 0),
            row.get("location") or "-",
        ]
        for row in data["low_stock_rows"]
    ]
    for row in stock_rows:
        ws_stock.append(row)
    _style_ws(ws_stock)

    # Apply formats and widths.
    for ws in wb.worksheets:
        for column in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
            ws.column_dimensions[get_column_letter(column[0].column)].width = min(max(max_len + 2, 12), 42)
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)) and cell.column != 1:
                    cell.alignment = Alignment(horizontal="right", vertical="top")
        if ws.title in {"Замовлення", "Товари", "Склад"}:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=4 if ws.title != "Склад" else 5).alignment = Alignment(horizontal="center", vertical="top")

    # Create bytes
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


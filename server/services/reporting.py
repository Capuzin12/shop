from __future__ import annotations

from datetime import datetime, timedelta
from io import BytesIO
import logging
from pathlib import Path
import traceback
from typing import Any
from xml.sax.saxutils import escape

from config import settings
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.pdfmetrics import registerFontFamily
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle, BaseDocTemplate, Frame, PageTemplate, PageBreak
from reportlab.graphics.shapes import Drawing, Rect, String as GString
from sqlalchemy import text
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import Session

logger = logging.getLogger(__name__)

ORDER_STATUS_LABELS = {
    "new": "Нове",
    "processing": "В обробці",
    "shipped": "Відправлено",
    "delivered": "Доставлено",
    "picked_up": "Забрано",
    "cancelled": "Скасовано",
    "refunded": "Повернено",
}

REPORT_FONT_CANDIDATES = [
    settings.report_font_path.strip(), str(Path("C:/Windows/Fonts/segoeui.ttf")),
    str(Path("C:/Windows/Fonts/arialuni.ttf")), str(Path("C:/Windows/Fonts/arial.ttf")),
    str(Path(__file__).resolve().parent / "fonts" / "DejaVuSans.ttf"),
    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    "/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf",
    "/usr/share/fonts/dejavu/DejaVuSans.ttf",
]

# ─── Колірна палітра та стилі для звіту ──────────────────────────────────────
C_DARK      = colors.HexColor("#0f172a")
C_SLATE     = colors.HexColor("#334155")
C_MUTED     = colors.HexColor("#64748b")
C_BORDER    = colors.HexColor("#e2e8f0")
C_AMBER     = colors.HexColor("#f59e0b")
C_EMERALD   = colors.HexColor("#10b981")
C_ROSE      = colors.HexColor("#f43f5e")
C_BLUE      = colors.HexColor("#3b82f6")
C_HEADER_BG = colors.HexColor("#1e293b")
C_ROW_ALT   = colors.HexColor("#f8fafc")
C_WHITE     = colors.white


def _safe_money(value) -> str:
    try:
        return f"{float(value or 0):,.2f}".replace(",", " ")
    except (TypeError, ValueError):
        return "0.00"

def _safe_int(value, default: int = 0) -> int:
    try:
        if value is None:
            return default
        return int(value)
    except (TypeError, ValueError):
        return default

def _status_label(status_value: str) -> str:
    return ORDER_STATUS_LABELS.get(str(status_value or "new"), str(status_value or "new"))

def _movement_label(movement_type: Any) -> str:
    labels = {
        "receipt": "Прийом", "sale": "Продаж", "return": "Повернення",
        "return_": "Повернення", "adjustment": "Коригування", "write_off": "Списання",
    }
    return labels.get(str(movement_type or "adjustment"), str(movement_type or "adjustment"))

def _month_window(reference: datetime, offset_months: int = 0) -> tuple[datetime, datetime]:
    year = reference.year + ((reference.month - 1 + offset_months) // 12)
    month = ((reference.month - 1 + offset_months) % 12) + 1
    start = datetime(year, month, 1)
    if month == 12:
        end = datetime(year + 1, 1, 1) - timedelta(days=1)
    else:
        end = datetime(year, month + 1, 1) - timedelta(days=1)
    return start, end

def _percent_change(current: float | int, previous: float | int) -> float:
    try:
        previous_value = float(previous or 0)
        if previous_value == 0: return 0.0
        return round(((float(current or 0) - previous_value) / previous_value) * 100.0, 2)
    except Exception: return 0.0

def _ratio(numerator: float | int, denominator: float | int) -> float:
    try:
        denominator_value = float(denominator or 0)
        if denominator_value == 0: return 0.0
        return round((float(numerator or 0) / denominator_value) * 100.0, 2)
    except Exception: return 0.0

def _month_label(dt_value: datetime) -> str:
    try: return dt_value.strftime("%m.%Y")
    except Exception: return "Період"

def _format_dt(value: Any) -> str:
    if value is None: return "-"
    if isinstance(value, datetime): return value.strftime("%d.%m.%Y %H:%M")
    text_value = str(value)
    if "T" in text_value: return text_value.replace("T", " ")[:16]
    return text_value[:16]

def _scalar_or_default(db: Session, statement, default=0, params: dict[str, Any] | None = None):
    try: return db.scalar(statement, params or {}) or default
    except SQLAlchemyError as error:
        logger.warning("report scalar query failed", extra={"error": str(error)})
        return default

def _rows_or_empty(db: Session, statement, params: dict[str, Any] | None = None):
    try: return db.execute(statement, params or {}).mappings().all()
    except SQLAlchemyError as error:
        logger.warning("report rows query failed", extra={"error": str(error)})
        return []

def collect_admin_report_data(db: Session) -> dict[str, Any]:
    generated_at = datetime.now()
    counts = {"categories": _scalar_or_default(db, text("SELECT COUNT(*) FROM categories"), 0), "products": _scalar_or_default(db, text("SELECT COUNT(*) FROM products"), 0), "orders": _scalar_or_default(db, text("SELECT COUNT(*) FROM orders"), 0), "users": _scalar_or_default(db, text("SELECT COUNT(*) FROM users"), 0)}
    current_start, current_end = _month_window(generated_at, 0)
    previous_start, previous_end = _month_window(generated_at, -1)
    current_start_date = current_start.date().isoformat()
    current_end_date = current_end.date().isoformat()
    previous_start_date = previous_start.date().isoformat()
    previous_end_date = previous_end.date().isoformat()

    inventory_rows = _rows_or_empty(db, text("""
                                             SELECT
                                                 i.id,
                                                 i.product_id,
                                                 COALESCE(i.quantity, 0) AS quantity,
                                                 COALESCE(i.min_quantity, 0) AS min_quantity,
                                                 COALESCE(i.min_quantity_alert, i.min_quantity, 0) AS threshold,
                                                 COALESCE(i.max_quantity, 0) AS max_quantity,
                                                 COALESCE(i.location, '-') AS location,
                                                 p.name AS product_name,
                                                 p.sku AS product_sku
                                             FROM inventory i
                                                      LEFT JOIN products p ON p.id = i.product_id
                                             ORDER BY COALESCE(i.quantity, 0) ASC, i.id ASC
                                             """))
    low_stock_rows = [row for row in inventory_rows if _safe_int(row["quantity"], 0) < _safe_int(row["threshold"], 0)]
    out_of_stock_rows = [row for row in inventory_rows if _safe_int(row["quantity"], 0) <= 0]
    total_stock_units = sum(max(_safe_int(row["quantity"], 0), 0) for row in inventory_rows)

    status_rows = _rows_or_empty(db, text("""
                                          SELECT COALESCE(status, 'new') AS status, COUNT(*) AS count
                                          FROM orders
                                          GROUP BY COALESCE(status, 'new')
                                          ORDER BY count DESC
                                          """))
    status_counts = [{"status": str(row["status"]), "label": _status_label(row["status"]), "count": _safe_int(row["count"], 0)} for row in status_rows]

    paid_revenue = _scalar_or_default(db, text("SELECT COALESCE(SUM(total), 0) FROM orders WHERE status IN ('delivered', 'picked_up')"), 0)
    current_month_revenue = _scalar_or_default(db, text("SELECT COALESCE(SUM(total), 0) FROM orders WHERE status IN ('delivered', 'picked_up') AND DATE(created_at) BETWEEN :start_date AND :end_date"), 0, {"start_date": current_start_date, "end_date": current_end_date})
    previous_month_revenue = _scalar_or_default(db, text("SELECT COALESCE(SUM(total), 0) FROM orders WHERE status IN ('delivered', 'picked_up') AND DATE(created_at) BETWEEN :start_date AND :end_date"), 0, {"start_date": previous_start_date, "end_date": previous_end_date})
    current_month_orders = _scalar_or_default(db, text("SELECT COUNT(*) FROM orders WHERE DATE(created_at) BETWEEN :start_date AND :end_date"), 0, {"start_date": current_start_date, "end_date": current_end_date})
    previous_month_orders = _scalar_or_default(db, text("SELECT COUNT(*) FROM orders WHERE DATE(created_at) BETWEEN :start_date AND :end_date"), 0, {"start_date": previous_start_date, "end_date": previous_end_date})
    current_month_fulfilled = _scalar_or_default(db, text("SELECT COUNT(*) FROM orders WHERE status IN ('delivered', 'picked_up') AND DATE(created_at) BETWEEN :start_date AND :end_date"), 0, {"start_date": current_start_date, "end_date": current_end_date})
    previous_month_fulfilled = _scalar_or_default(db, text("SELECT COUNT(*) FROM orders WHERE status IN ('delivered', 'picked_up') AND DATE(created_at) BETWEEN :start_date AND :end_date"), 0, {"start_date": previous_start_date, "end_date": previous_end_date})
    current_month_users = _scalar_or_default(db, text("SELECT COUNT(*) FROM users WHERE DATE(created_at) BETWEEN :start_date AND :end_date"), 0, {"start_date": current_start_date, "end_date": current_end_date})
    previous_month_users = _scalar_or_default(db, text("SELECT COUNT(*) FROM users WHERE DATE(created_at) BETWEEN :start_date AND :end_date"), 0, {"start_date": previous_start_date, "end_date": previous_end_date})
    avg_order_value = float(paid_revenue or 0) / float(current_month_fulfilled or 0) if int(current_month_fulfilled or 0) > 0 else 0.0
    prev_avg_order_value = float(previous_month_revenue or 0) / float(previous_month_fulfilled or 0) if int(previous_month_fulfilled or 0) > 0 else 0.0

    comparison = {
        "current_month": {
            "label": _month_label(current_start),
            "revenue": float(current_month_revenue or 0),
            "orders": int(current_month_orders or 0),
            "fulfilled_orders": int(current_month_fulfilled or 0),
            "avg_order_value": float(avg_order_value),
            "conversion_rate": _ratio(current_month_fulfilled, current_month_orders),
            "new_users": int(current_month_users or 0),
        },
        "previous_month": {
            "label": _month_label(previous_start),
            "revenue": float(previous_month_revenue or 0),
            "orders": int(previous_month_orders or 0),
            "fulfilled_orders": int(previous_month_fulfilled or 0),
            "avg_order_value": float(prev_avg_order_value),
            "conversion_rate": _ratio(previous_month_fulfilled, previous_month_orders),
            "new_users": int(previous_month_users or 0),
        },
    }

    latest_orders = _rows_or_empty(db, text("SELECT id, user_id, contact_name, contact_phone, contact_email, delivery_city, delivery_address, COALESCE(status, 'new') AS status, COALESCE(total, 0) AS total, created_at FROM orders ORDER BY created_at DESC LIMIT 12"))
    top_products = _rows_or_empty(db, text("SELECT id, name, sku, COALESCE(price, 0) AS price, CASE WHEN is_active IS TRUE THEN 1 ELSE 0 END AS is_active FROM products ORDER BY COALESCE(price, 0) DESC, id DESC LIMIT 10"))

    movement_since = (datetime.utcnow() - timedelta(days=30)).date().isoformat()
    inventory_movement_summary = _rows_or_empty(db, text("""
                                                         SELECT COALESCE(type, 'adjustment') AS type, COUNT(*) AS total_movements, COALESCE(SUM(quantity), 0) AS net_change, COALESCE(SUM(CASE WHEN quantity > 0 THEN quantity ELSE 0 END), 0) AS incoming_quantity, COALESCE(SUM(CASE WHEN quantity < 0 THEN -quantity ELSE 0 END), 0) AS outgoing_quantity
                                                         FROM inventory_movements WHERE DATE(created_at) >= :since GROUP BY COALESCE(type, 'adjustment') ORDER BY total_movements DESC, type ASC
                                                         """), {"since": movement_since})

    recent_inventory_movements = _rows_or_empty(db, text("""
                                                         SELECT m.id, m.type, m.quantity, m.quantity_before, m.quantity_after, m.note, m.created_at, p.name AS product_name, p.sku AS product_sku, u.first_name AS created_by_first_name, u.last_name AS created_by_last_name, so.invoice_number AS supply_invoice_number, so.id AS supply_order_id, o.id AS order_id
                                                         FROM inventory_movements m JOIN products p ON p.id = m.product_id LEFT JOIN users u ON u.id = m.created_by LEFT JOIN supply_orders so ON so.id = m.supply_order_id LEFT JOIN orders o ON o.id = m.order_id
                                                         WHERE DATE(m.created_at) >= :since ORDER BY m.created_at DESC, m.id DESC LIMIT 12
                                                         """), {"since": movement_since})

    def _movement_source_label(row: dict[str, Any]) -> str:
        if row.get("order_id") is not None: return f"Замовлення #{row['order_id']}"
        if row.get("supply_order_id") is not None:
            invoice = row.get("supply_invoice_number")
            return f"Поставка {invoice}" if invoice else f"Поставка #{row['supply_order_id']}"
        return "Ручне коригування"

    inventory_movement_rows = []
    for row in recent_inventory_movements:
        created_by_name = f"{row['created_by_first_name'] or ''} {row['created_by_last_name'] or ''}".strip() or "Система"
        inventory_movement_rows.append({"id": _safe_int(row.get("id"), 0), "type": str(row.get("type") or "adjustment"), "quantity": _safe_int(row.get("quantity"), 0), "quantity_before": _safe_int(row.get("quantity_before"), 0), "quantity_after": _safe_int(row.get("quantity_after"), 0), "note": row.get("note"), "created_at": row.get("created_at"), "product_name": row.get("product_name"), "product_sku": row.get("product_sku"), "created_by_name": created_by_name, "source": _movement_source_label(row)})

    try:
        days = 30
        start_date = (datetime.utcnow() - timedelta(days=days)).date().isoformat()
        revenue_trend = db.execute(text("SELECT DATE(created_at) as day, COALESCE(SUM(total), 0) as revenue, COUNT(*) as orders_count FROM orders WHERE status IN ('delivered', 'picked_up') AND DATE(created_at) >= :since GROUP BY DATE(created_at) ORDER BY DATE(created_at) ASC"), {"since": start_date}).mappings().all()
    except Exception:
        revenue_trend = _rows_or_empty(db, text("SELECT DATE(created_at) as day, COALESCE(SUM(total), 0) as revenue, COUNT(*) as orders_count FROM orders WHERE status IN ('delivered', 'picked_up') GROUP BY DATE(created_at) ORDER BY DATE(created_at) ASC"))

    top_products_revenue = _rows_or_empty(db, text("""
                                                   SELECT p.id AS product_id, p.name, p.sku, SUM(oi.quantity) AS total_qty, SUM(oi.quantity * oi.unit_price) AS total_revenue, COUNT(DISTINCT o.id) AS orders_count
                                                   FROM order_items oi JOIN orders o ON o.id = oi.order_id JOIN products p ON p.id = oi.product_id
                                                   WHERE o.status IN ('delivered', 'picked_up') GROUP BY p.id, p.name, p.sku ORDER BY total_revenue DESC LIMIT 10
                                                   """))

    top_customers = _rows_or_empty(db, text("""
                                            SELECT u.id AS user_id, COALESCE(NULLIF(TRIM(COALESCE(u.first_name, '') || ' ' || COALESCE(u.last_name, '')), ''), u.email) AS name, COUNT(o.id) AS orders_count, COALESCE(SUM(o.total), 0) AS total_spent, CASE WHEN COUNT(o.id) > 0 THEN COALESCE(SUM(o.total), 0) / COUNT(o.id) ELSE 0 END AS avg_order_value
                                            FROM users u JOIN orders o ON o.user_id = u.id
                                            WHERE o.status IN ('delivered', 'picked_up') AND DATE(o.created_at) >= :since GROUP BY u.id, u.first_name, u.last_name, u.email ORDER BY total_spent DESC LIMIT 10
                                            """), {"since": movement_since})

    problematic_products = _rows_or_empty(db, text("""
                                                   SELECT p.id AS product_id, p.name, p.sku, COUNT(*) AS low_stock_hits, MAX(m.created_at) AS last_low_stock_at, MIN(m.quantity_after) AS lowest_quantity_after, COALESCE(i.quantity, 0) AS current_quantity, COALESCE(i.min_quantity_alert, i.min_quantity, 0) AS threshold
                                                   FROM inventory_movements m JOIN inventory i ON i.product_id = m.product_id JOIN products p ON p.id = m.product_id
                                                   WHERE DATE(m.created_at) >= :since AND m.quantity_after < COALESCE(i.min_quantity_alert, i.min_quantity, 0) GROUP BY p.id, p.name, p.sku, i.quantity, i.min_quantity_alert, i.min_quantity ORDER BY low_stock_hits DESC, last_low_stock_at DESC LIMIT 10
                                                   """), {"since": movement_since})

    inventory_anomalies = _rows_or_empty(db, text("""
                                                  SELECT p.id AS product_id, p.name, p.sku, COUNT(*) AS anomaly_count, SUM(CASE WHEN m.type = 'write_off' THEN 1 ELSE 0 END) AS write_off_count, SUM(CASE WHEN m.type = 'adjustment' THEN 1 ELSE 0 END) AS adjustment_count, SUM(CASE WHEN ABS(m.quantity) >= 20 THEN 1 ELSE 0 END) AS large_delta_count, MAX(m.created_at) AS last_anomaly_at, MAX(ABS(m.quantity)) AS max_delta
                                                  FROM inventory_movements m JOIN products p ON p.id = m.product_id
                                                  WHERE DATE(m.created_at) >= :since AND (m.type IN ('adjustment', 'write_off') OR ABS(m.quantity) >= 20) GROUP BY p.id, p.name, p.sku ORDER BY anomaly_count DESC, large_delta_count DESC, last_anomaly_at DESC LIMIT 10
                                                  """), {"since": movement_since})

    customer_stats_total = _scalar_or_default(db, text("SELECT COUNT(DISTINCT user_id) FROM orders WHERE status IN ('delivered', 'picked_up')"), default=0)

    data = {
        "generated_at": generated_at,
        "counts": counts,
        "comparison": comparison,
        "current_month_revenue": float(current_month_revenue or 0),
        "previous_month_revenue": float(previous_month_revenue or 0),
        "current_month_orders": int(current_month_orders or 0),
        "previous_month_orders": int(previous_month_orders or 0),
        "current_month_fulfilled": int(current_month_fulfilled or 0),
        "previous_month_fulfilled": int(previous_month_fulfilled or 0),
        "current_month_users": int(current_month_users or 0),
        "previous_month_users": int(previous_month_users or 0),
        "avg_order_value": float(avg_order_value),
        "prev_avg_order_value": float(prev_avg_order_value),
        "total_stock_units": total_stock_units,
        "low_stock_rows": low_stock_rows,
        "out_of_stock_rows": out_of_stock_rows,
        "status_counts": status_counts,
        "paid_revenue": paid_revenue,
        "latest_orders": latest_orders,
        "top_products": top_products,
        "revenue_trend": revenue_trend,
        "top_products_revenue": top_products_revenue,
        "top_customers": top_customers,
        "problematic_products": problematic_products,
        "inventory_anomalies": inventory_anomalies,
        "customer_stats_total_orders_customers": customer_stats_total,
        "inventory_movement_summary": inventory_movement_summary,
        "inventory_movement_rows": inventory_movement_rows,
    }

    return data


def _resolve_pdf_font() -> str:
    for candidate in REPORT_FONT_CANDIDATES:
        if not candidate:
            continue
        path = Path(candidate)
        if not path.exists() or not path.is_file():
            continue

        font_name = f"BuildShop-{path.stem}"
        try:
            if font_name not in pdfmetrics.getRegisteredFontNames():
                font_obj = TTFont(font_name, str(path))
                pdfmetrics.registerFont(font_obj)
                try:
                    registerFontFamily(font_name, normal=font_name, bold=font_name, italic=font_name, boldItalic=font_name)
                except Exception:
                    pass
            logger.info("PDF report font selected", extra={"font_name": font_name, "font_path": str(path)})
            return font_name
        except Exception as error:
            logger.warning("PDF font candidate failed to load", extra={"font_path": str(path), "error": str(error)})
            continue

    logger.warning("No valid TTF fonts found. Falling back to default Helvetica.")
    return "Helvetica"

def _kpi_card(label, value, font_bold, font_reg, unit="", accent=C_AMBER, width=42*mm, height=26*mm):
    d = Drawing(width, height)
    d.add(Rect(0, 0, width, height, rx=4, ry=4, fillColor=colors.HexColor("#f8fafc"), strokeColor=C_BORDER, strokeWidth=0.8))
    d.add(Rect(0, height-3, width, 3, rx=1, ry=1, fillColor=accent, strokeColor=None))
    d.add(GString(width/2, height/2 + 2, str(value), fontName=font_bold, fontSize=15, textAnchor="middle", fillColor=C_DARK))
    if unit: d.add(GString(width/2, height/2 - 7, unit, fontName=font_reg, fontSize=6.5, textAnchor="middle", fillColor=C_MUTED))
    d.add(GString(width/2, 5, label, fontName=font_reg, fontSize=6.5, textAnchor="middle", fillColor=C_MUTED))
    return d

def _section_header(number, title, font_bold, accent=C_AMBER, width=174*mm):
    d = Drawing(width, 10*mm)
    d.add(Rect(0, 0, width, 10*mm, rx=3, ry=3, fillColor=colors.HexColor("#f1f5f9"), strokeColor=C_BORDER, strokeWidth=0.5))
    d.add(Rect(0, 0, 4*mm, 10*mm, fillColor=accent, strokeColor=None))
    d.add(GString(8*mm, 3.5*mm, f"{number}. {title}", fontName=font_bold, fontSize=11, fillColor=C_DARK))
    return d

def _simple_bar(value, max_value, width=60*mm, height=4*mm, color=C_BLUE):
    """Малює просту горизонтальну лінію у комірці таблиці (без тексту)"""
    d = Drawing(width, height)
    # Background track (сірий)
    d.add(Rect(0, 0, width, height, rx=1.5, ry=1.5, fillColor=colors.HexColor("#e2e8f0"), strokeColor=None))
    # Active bar (синій)
    if max_value > 0 and value > 0:
        bar_w = (value / max_value) * width
        d.add(Rect(0, 0, max(bar_w, 2), height, rx=1.5, ry=1.5, fillColor=color, strokeColor=None))
    return d


def build_admin_report_pdf(data: dict[str, Any]) -> bytes:
    """Глобальна обгортка: перехоплює будь-яке падіння і повертає PDF з логом помилки"""
    try:
        return _build_admin_report_pdf_internal(data)
    except Exception as e:
        logger.error(f"PDF Generation failed: {e}\n{traceback.format_exc()}")
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        err_text = escape(traceback.format_exc()).replace('\n', '<br/>')
        story = [
            Paragraph("Помилка генерації PDF", styles["Heading1"]),
            Paragraph("Під час створення звіту сталася помилка. Сервер уникнув падіння (500 Error), щоб показати вам цей лог:", styles["Normal"]),
            Spacer(1, 5*mm),
            Paragraph(f"<font color='red'>{err_text}</font>", styles["Normal"])
        ]
        doc.build(story)
        return buffer.getvalue()


def _build_admin_report_pdf_internal(data: dict[str, Any]) -> bytes:
    font_name = _resolve_pdf_font()
    font_bold = font_name

    buffer = BytesIO()
    MARGIN = 18 * mm
    HEADER_H = 28 * mm
    FOOTER_H = 10 * mm
    PAGE_W, PAGE_H = A4

    doc = BaseDocTemplate(
        buffer, pagesize=A4,
        leftMargin=MARGIN, rightMargin=MARGIN,
        topMargin=HEADER_H + 5*mm, bottomMargin=FOOTER_H + 5*mm,
        title="BuildShop Адміністративний звіт", author="BuildShop",
    )

    def on_first_page(canvas, doc):
        canvas.saveState()
        canvas.setFillColor(C_HEADER_BG)
        canvas.rect(0, PAGE_H - HEADER_H, PAGE_W, HEADER_H, fill=1, stroke=0)
        canvas.setFillColor(C_AMBER)
        canvas.rect(0, PAGE_H - HEADER_H - 2.5*mm, PAGE_W, 2.5*mm, fill=1, stroke=0)
        canvas.setFillColor(C_AMBER)
        canvas.setFont(font_bold, 22)
        canvas.drawString(MARGIN, PAGE_H - 17*mm, "BUILDSHOP")
        canvas.setFillColor(colors.HexColor("#94a3b8"))
        canvas.setFont(font_name, 9)
        canvas.drawString(MARGIN, PAGE_H - 24*mm, "Адміністративний звіт")
        canvas.setFont(font_name, 8)
        canvas.drawRightString(PAGE_W - MARGIN, PAGE_H - 24*mm, data['generated_at'].strftime("Сформовано: %d.%m.%Y %H:%M"))
        _draw_footer(canvas, doc)
        canvas.restoreState()

    def on_later_pages(canvas, doc):
        canvas.saveState()
        canvas.setFillColor(C_HEADER_BG)
        canvas.rect(0, PAGE_H - 10*mm, PAGE_W, 10*mm, fill=1, stroke=0)
        canvas.setFillColor(C_AMBER)
        canvas.rect(0, PAGE_H - 10*mm - 1.5*mm, PAGE_W, 1.5*mm, fill=1, stroke=0)
        canvas.setFillColor(C_AMBER)
        canvas.setFont(font_bold, 9)
        canvas.drawString(MARGIN, PAGE_H - 7*mm, "BUILDSHOP")
        canvas.setFillColor(colors.HexColor("#94a3b8"))
        canvas.setFont(font_name, 8)
        canvas.drawRightString(PAGE_W - MARGIN, PAGE_H - 7*mm, f"Адміністративний звіт · {data['generated_at'].strftime('%d.%m.%Y')}")
        _draw_footer(canvas, doc)
        canvas.restoreState()

    def _draw_footer(canvas, doc):
        canvas.setFillColor(C_BORDER)
        canvas.rect(0, 0, PAGE_W, FOOTER_H, fill=1, stroke=0)
        canvas.setFillColor(C_MUTED)
        canvas.setFont(font_name, 7)
        canvas.drawString(MARGIN, 3.5*mm, "© 2026 BuildShop — Конфіденційно")
        canvas.drawRightString(PAGE_W - MARGIN, 3.5*mm, f"Сторінка {doc.page}")

    first_frame = Frame(MARGIN, FOOTER_H + 5*mm, PAGE_W - 2*MARGIN, PAGE_H - HEADER_H - FOOTER_H - 10*mm, id="first")
    later_frame = Frame(MARGIN, FOOTER_H + 5*mm, PAGE_W - 2*MARGIN, PAGE_H - 10*mm - FOOTER_H - 10*mm, id="later")
    doc.addPageTemplates([
        PageTemplate(id="First", frames=[first_frame], onPage=on_first_page),
        PageTemplate(id="Later", frames=[later_frame], onPage=on_later_pages),
    ])

    story = []
    W = PAGE_W - 2*MARGIN

    def S(name, size=8, color=C_SLATE, align=TA_LEFT, bold=False):
        return ParagraphStyle(name, fontName=font_bold if bold else font_name, fontSize=size, textColor=color, alignment=align, leading=size*1.35)

    sBody = S("Body")
    sBodyC = S("BodyC", align=TA_CENTER)
    sBodyR = S("BodyR", align=TA_RIGHT)
    sTH = S("TH", color=C_WHITE, bold=True, align=TA_CENTER)
    sTHL = S("THL", color=C_WHITE, bold=True, align=TA_LEFT)
    sNote = S("Note", size=7.5, color=C_MUTED)
    sGreen = S("Green", color=C_EMERALD, bold=True, align=TA_CENTER)
    sRed = S("Red", color=C_ROSE, bold=True, align=TA_CENTER)
    sAmber = S("Amber", color=C_AMBER, bold=True, align=TA_CENTER)

    # ЕКРАНУВАННЯ: Завжди очищуємо текст перед передачею в Paragraph
    def p(text, style):
        safe_text = escape(str(text if text is not None else "—")).replace("\n", "<br/>")
        return Paragraph(safe_text, style)

    def make_table(rows, col_widths):
        t = Table(rows, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), C_HEADER_BG),
            ("TEXTCOLOR", (0, 0), (-1, 0), C_WHITE),
            ("FONTNAME", (0, 0), (-1, 0), font_bold),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
            ("LEADING", (0, 0), (-1, -1), 10),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [C_WHITE, C_ROW_ALT]),
            ("GRID", (0, 0), (-1, -1), 0.5, C_BORDER),
            ("LINEBELOW", (0, 0), (-1, 0), 1.2, C_HEADER_BG),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("FONTNAME", (0, 1), (-1, -1), font_name),
        ]))
        return t

    counts = data.get("counts", {})
    comp = data.get("comparison", {})
    cur = comp.get("current_month", {})
    prev = comp.get("previous_month", {})

    # 1. KPI Cards
    story.append(Spacer(1, 3*mm))
    divider = Drawing(W, 0.5*mm)
    divider.add(Rect(0, 0, W, 0.5, fillColor=C_AMBER, strokeColor=None))
    story.append(divider)
    story.append(Spacer(1, 4*mm))

    card_w = (W - 4*mm) / 5
    cards = [
        ("Категорії", str(counts.get("categories", 0)), "всього", C_BLUE),
        ("Товари", str(counts.get("products", 0)), "позицій", C_AMBER),
        ("Замовлення", str(counts.get("orders", 0)), "всього", C_EMERALD),
        ("Користувачі", str(counts.get("users", 0)), "акаунтів", C_SLATE),
        ("Виторг", _safe_money(data.get("paid_revenue")), "грн", C_ROSE),
    ]
    card_row = [[_kpi_card(lbl, val, font_bold, font_name, unt, acc, card_w - 1*mm, 28*mm) for lbl, val, unt, acc in cards]]
    ct = Table(card_row, colWidths=[card_w]*5, hAlign="CENTER")
    ct.setStyle(TableStyle([("LEFTPADDING", (0,0), (-1,-1), 0), ("RIGHTPADDING", (0,0), (-1,-1), 0)]))
    story.append(ct)
    story.append(Spacer(1, 6*mm))

    # 2. Порівняння
    story.append(_section_header("1", "Порівняння з попереднім місяцем", font_bold, width=W))
    story.append(Spacer(1, 3*mm))

    comp_rows = [[p("Метрика", sTHL), p(cur.get("label") or "Поточний", sTH), p(prev.get("label") or "Попередній", sTH), p("Зміна", sTH)]]
    metrics = [
        ("Виторг, грн", _safe_money(cur.get("revenue")), _safe_money(prev.get("revenue")), cur.get("revenue", 0), prev.get("revenue", 0)),
        ("Замовлення", str(cur.get("orders", 0)), str(prev.get("orders", 0)), cur.get("orders", 0), prev.get("orders", 0)),
        ("Виконані", str(cur.get("fulfilled_orders", 0)), str(prev.get("fulfilled_orders", 0)), cur.get("fulfilled_orders", 0), prev.get("fulfilled_orders", 0)),
        ("Середній чек, грн", _safe_money(cur.get("avg_order_value")), _safe_money(prev.get("avg_order_value")), cur.get("avg_order_value", 0), prev.get("avg_order_value", 0)),
        ("Нові користувачі", str(cur.get("new_users", 0)), str(prev.get("new_users", 0)), cur.get("new_users", 0), prev.get("new_users", 0)),
    ]
    for label, cv_str, pv_str, cv, pv in metrics:
        delta = _percent_change(cv, pv)
        d_str = f"▲ {abs(delta):.1f}%" if delta >= 0 else f"▼ {abs(delta):.1f}%"
        st = sGreen if delta >= 0 else sRed
        if pv == 0: d_str, st = "—", sBodyC
        comp_rows.append([p(label, sBody), p(cv_str, sBodyC), p(pv_str, sBodyC), p(d_str, st)])
    story.append(make_table(comp_rows, [W*0.32, W*0.22, W*0.22, W*0.24]))
    story.append(Spacer(1, 6*mm))

    # 3. Статуси замовлень - ТЕПЕР ІДЕАЛЬНО ВИРІВНЯНО
    story.append(_section_header("2", "Статуси замовлень", font_bold, width=W))
    story.append(Spacer(1, 3*mm))
    status_rows_data = data.get("status_counts", [])
    if status_rows_data:
        max_cnt = max((int(r.get("count", 0)) for r in status_rows_data), default=1)
        sr = [[p("Статус", sTHL), p("Кількість", sTHC if 'sTHC' in locals() else sTH), p("Графік", sTHL)]]

        sorted_status = sorted(status_rows_data, key=lambda x: int(x.get("count", 0)), reverse=True)

        for row in sorted_status:
            label = row.get("label", "")
            cnt = int(row.get("count", 0))
            bar_drawing = _simple_bar(cnt, max_cnt, width=W*0.45, height=5*mm, color=C_BLUE)
            sr.append([p(label, sBody), p(str(cnt), sBodyC), bar_drawing])

        t_status = make_table(sr, [W*0.35, W*0.15, W*0.50])
        story.append(t_status)
    else:
        story.append(p("Немає даних.", sNote))
    story.append(Spacer(1, 6*mm))

    # 4. Останні замовлення
    story.append(_section_header("3", "Останні замовлення", font_bold, width=W))
    story.append(Spacer(1, 3*mm))
    if data.get("latest_orders"):
        r4 = [[p("№", sTH), p("Дата", sTH), p("Клієнт", sTHL), p("Статус", sTH), p("Сума, грн", sTH)]]
        for row in data["latest_orders"][:8]:
            r4.append([p(f"#{row['id']}", sBodyC), p(_format_dt(row.get("created_at")), sBodyC), p(row.get("contact_name") or "—", sBody), p(_status_label(row.get("status")), sBodyC), p(_safe_money(row.get("total")), sBodyR)])
        story.append(make_table(r4, [W*0.07, W*0.20, W*0.30, W*0.18, W*0.25]))
    else:
        story.append(p("Замовлень ще немає.", sNote))
    story.append(Spacer(1, 6*mm))

    # 5. Топ товари
    story.append(_section_header("4", "Топ товарів за виторгом", font_bold, width=W))
    story.append(Spacer(1, 3*mm))
    if data.get("top_products_revenue"):
        r5 = [[p("#", sTH), p("Назва / SKU", sTHL), p("Продано", sTH), p("Замовлень", sTH), p("Виторг, грн", sTH)]]
        for i, row in enumerate(data["top_products_revenue"][:8], 1):
            name_safe = escape(str(row.get('name') or '—'))
            sku_safe = escape(str(row.get('sku') or ''))
            r5.append([
                p(str(i), sBodyC),
                Paragraph(f"<b>{name_safe}</b><br/><font size='6.5' color='#64748b'>{sku_safe}</font>", sBody),
                p(str(row.get("total_qty", 0)), sBodyC),
                p(str(row.get("orders_count", 0)), sBodyC),
                p(_safe_money(row.get("total_revenue")), sBodyR)
            ])
        story.append(make_table(r5, [W*0.04, W*0.46, W*0.14, W*0.16, W*0.20]))
    else:
        story.append(p("Немає даних.", sNote))

    story.append(PageBreak())

    # 6. Склад і Рух
    story.append(_section_header("5", "Склад і критичні залишки", font_bold, width=W))
    story.append(Spacer(1, 3*mm))
    if data.get("low_stock_rows"):
        r6 = [[p("Товар", sTHL), p("SKU", sTH), p("Залишок", sTH), p("Поріг", sTH), p("Локація", sTH)]]
        for row in data["low_stock_rows"][:15]:
            qty = _safe_int(row.get("quantity"))
            st = sRed if qty <= 0 else sAmber
            r6.append([p(row.get("product_name") or "—", sBody), p(row.get("product_sku") or "—", sBodyC), p(str(qty), st), p(str(row.get("threshold", 0)), sBodyC), p(row.get("location") or "—", sBodyC)])
        story.append(make_table(r6, [W*0.38, W*0.15, W*0.12, W*0.12, W*0.23]))
    else:
        story.append(p("Критичних залишків немає.", sNote))

    doc.build(story)
    return buffer.getvalue()


def _style_ws(ws, header_row: int = 1):
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row in ws.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = ws[2][0].coordinate if ws.max_row > header_row else "A2"
    ws.auto_filter.ref = ws.dimensions
    for column_cells in ws.columns:
        column_letter = column_cells[0].column_letter
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 42)


def build_admin_report_xlsx(data: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Підсумок"

    ws_summary["A1"] = "BuildShop — Адміністративний звіт"
    ws_summary["A1"].font = Font(size=16, bold=True, color="0F172A")
    ws_summary["A2"] = f"Дата формування: {data['generated_at'].strftime('%d.%m.%Y %H:%M')}"
    ws_summary["A3"] = "Звіт містить ключові показники, статуси замовлень, топ-товари та ризики по складу."
    ws_summary["A5"] = "Ключові показники"
    ws_summary["A5"].font = Font(size=12, bold=True)

    summary_rows = [["Показник", "Значення"], ["Категорії", data["counts"]["categories"]], ["Товари", data["counts"]["products"]], ["Замовлення", data["counts"]["orders"]], ["Користувачі", data["counts"]["users"]], ["Одиниць товару на складі", data["total_stock_units"]], ["Низький запас", len(data["low_stock_rows"])], ["Немає в наявності", len(data["out_of_stock_rows"])], ["Виторг (доставлено/забрано), грн", _safe_money(data["paid_revenue"])]]
    for row in summary_rows:
        ws_summary.append(row)

    start_row = len(summary_rows) + 3
    ws_summary.cell(row=start_row, column=1, value="Статуси замовлень").font = Font(size=12, bold=True)
    ws_summary.cell(row=start_row + 1, column=1, value="Статус")
    ws_summary.cell(row=start_row + 1, column=2, value="Кількість")
    for index, row in enumerate(data["status_counts"], start=start_row + 2):
        ws_summary.cell(row=index, column=1, value=row["label"])
        ws_summary.cell(row=index, column=2, value=row["count"])
    _style_ws(ws_summary)

    ws_orders = wb.create_sheet("Замовлення")
    order_rows = [["№", "Дата", "Клієнт", "Статус", "Сума, грн"]] + [[f"#{row['id']}", _format_dt(row.get("created_at")), row.get("contact_name") or f"user #{row.get('user_id')}", _status_label(row.get("status")), float(row.get("total") or 0)] for row in data["latest_orders"]]
    for row in order_rows:
        ws_orders.append(row)
    _style_ws(ws_orders)

    ws_products = wb.create_sheet("Товари")
    product_rows = [["Товар", "SKU", "Ціна, грн", "Активний"]] + [[row.get("name") or "-", row.get("sku") or "-", float(row.get("price") or 0), "Так" if _safe_int(row.get("is_active"), 0) else "Ні"] for row in data["top_products"]]
    for row in product_rows:
        ws_products.append(row)
    _style_ws(ws_products)

    ws_stock = wb.create_sheet("Склад")
    stock_rows = [["Товар", "SKU", "К-сть", "Поріг", "Локація"]] + [[row.get("product_name") or "Невідомо", row.get("product_sku") or "-", _safe_int(row.get("quantity"), 0), _safe_int(row.get("threshold"), 0), row.get("location") or "-"] for row in data["low_stock_rows"]]
    for row in stock_rows:
        ws_stock.append(row)
    _style_ws(ws_stock)

    ws_comparison = wb.create_sheet("Порівняння")
    comparison = data.get("comparison", {})
    current_month = comparison.get("current_month", {})
    previous_month = comparison.get("previous_month", {})
    comparison_rows = [
        ["Метрика", current_month.get("label") or "Поточний", previous_month.get("label") or "Попередній", "Зміна, %"],
        ["Виторг, грн", float(current_month.get("revenue") or 0), float(previous_month.get("revenue") or 0), _percent_change(current_month.get("revenue"), previous_month.get("revenue"))],
        ["Замовлення", _safe_int(current_month.get("orders"), 0), _safe_int(previous_month.get("orders"), 0), _percent_change(current_month.get("orders"), previous_month.get("orders"))],
        ["Виконані", _safe_int(current_month.get("fulfilled_orders"), 0), _safe_int(previous_month.get("fulfilled_orders"), 0), _percent_change(current_month.get("fulfilled_orders"), previous_month.get("fulfilled_orders"))],
        ["Середній чек, грн", float(current_month.get("avg_order_value") or 0), float(previous_month.get("avg_order_value") or 0), _percent_change(current_month.get("avg_order_value"), previous_month.get("avg_order_value"))],
        ["Конверсія, %", float(current_month.get("conversion_rate") or 0), float(previous_month.get("conversion_rate") or 0), _percent_change(current_month.get("conversion_rate"), previous_month.get("conversion_rate"))],
        ["Нові користувачі", _safe_int(current_month.get("new_users"), 0), _safe_int(previous_month.get("new_users"), 0), _percent_change(current_month.get("new_users"), previous_month.get("new_users"))],
    ]
    for row in comparison_rows:
        ws_comparison.append(row)
    _style_ws(ws_comparison)

    ws_movement = wb.create_sheet("Рух складу")
    movement_summary_rows = [["Тип", "К-сть рухів", "Надійшло", "Видано", "Чистий рух"]] + [[
        _movement_label(row.get("type")),
        _safe_int(row.get("total_movements"), 0),
        _safe_int(row.get("incoming_quantity"), 0),
        _safe_int(row.get("outgoing_quantity"), 0),
        _safe_int(row.get("net_change"), 0),
    ] for row in data.get("inventory_movement_summary", [])]
    for row in movement_summary_rows:
        ws_movement.append(row)
    _style_ws(ws_movement)

    try:
        from openpyxl.chart import BarChart, Reference
        if ws_movement.max_row >= 2:
            chart = BarChart()
            chart.type = "bar"
            chart.style = 10
            chart.title = "Рух складу за типами"
            chart.y_axis.title = "Тип"
            chart.x_axis.title = "К-сть рухів"
            data_ref = Reference(ws_movement, min_col=2, min_row=1, max_row=min(ws_movement.max_row, 1 + len(data.get("inventory_movement_summary", []))))
            cats_ref = Reference(ws_movement, min_col=1, min_row=2, max_row=min(ws_movement.max_row, 1 + len(data.get("inventory_movement_summary", []))))
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.height = 7
            chart.width = 10
            ws_movement.add_chart(chart, "H2")
    except Exception:
        logger.warning("Failed to add inventory movement chart to Excel report")

    start_row = len(movement_summary_rows) + 3
    ws_movement.cell(row=start_row, column=1, value="Останні рухи").font = Font(size=12, bold=True)
    recent_headers = ["Дата", "Товар", "Джерело", "Тип", "Зміна", "Стало"]
    for idx, header in enumerate(recent_headers, start=1):
        ws_movement.cell(row=start_row + 1, column=idx, value=header)
    for row_index, row in enumerate(data.get("inventory_movement_rows", []), start=start_row + 2):
        ws_movement.cell(row=row_index, column=1, value=_format_dt(row.get("created_at")))
        ws_movement.cell(row=row_index, column=2, value=f"{row.get('product_name') or '-'} ({row.get('product_sku') or '-'})")
        ws_movement.cell(row=row_index, column=3, value=row.get("source") or "-")
        ws_movement.cell(row=row_index, column=4, value=_movement_label(row.get("type")))
        ws_movement.cell(row=row_index, column=5, value=_safe_int(row.get("quantity"), 0))
        ws_movement.cell(row=row_index, column=6, value=_safe_int(row.get("quantity_after"), 0))
    _style_ws(ws_movement, header_row=start_row + 1)

    ws_revenue = wb.create_sheet("Динаміка виторгу")
    revenue_rows = [["Дата", "Виторг, грн", "Замовлень"]] + [
        [row["day"].strftime("%Y-%m-%d") if hasattr(row["day"], "strftime") else str(row["day"]), float(row["revenue"] or 0), int(row["orders_count"] or 0)]
        for row in data.get("revenue_trend", [])
    ]
    for row in revenue_rows:
        ws_revenue.append(row)
    _style_ws(ws_revenue)

    try:
        from openpyxl.chart import LineChart, Reference
        chart = LineChart()
        chart.title = "Динаміка виторгу"
        chart.style = 10
        chart.y_axis.title = "Виторг, грн"
        chart.x_axis.title = "Дата"
        if ws_revenue.max_row >= 2:
            data_ref = Reference(ws_revenue, min_col=2, min_row=1, max_row=ws_revenue.max_row)
            chart.add_data(data_ref, titles_from_data=True)
            cats = Reference(ws_revenue, min_col=1, min_row=2, max_row=ws_revenue.max_row)
            chart.set_categories(cats)
            ws_revenue.add_chart(chart, "E2")
    except Exception:
        logger.warning("Failed to add revenue chart to Excel report")

    ws_top = wb.create_sheet("Топ товари")
    top_rows = [["Товар", "SKU", "Продано (шт)", "Виторг, грн", "Замовлень"]] + [
        [r.get("name") or "-", r.get("sku") or "-", int(r.get("total_qty") or 0), float(r.get("total_revenue") or 0), int(r.get("orders_count") or 0)]
        for r in data.get("top_products_revenue", [])
    ]
    for row in top_rows:
        ws_top.append(row)
    _style_ws(ws_top)

    ws_customers = wb.create_sheet("Клієнти")
    customer_rows = [["Клієнт", "Замовлень", "Витрачено, грн", "Сер. чек, грн"]] + [
        [row.get("name") or "-", _safe_int(row.get("orders_count"), 0), float(row.get("total_spent") or 0), float(row.get("avg_order_value") or 0)]
        for row in data.get("top_customers", [])
    ]
    if not data.get("top_customers"):
        customer_rows.append(["Немає даних", 0, 0, 0])
    for row in customer_rows:
        ws_customers.append(row)
    _style_ws(ws_customers)

    ws_problematic = wb.create_sheet("Проблемні товари")
    problematic_rows = [["Товар", "SKU", "Спрацювань", "Найнижче", "Поточний", "Поріг"]] + [
        [row.get("name") or "-", row.get("sku") or "-", _safe_int(row.get("low_stock_hits"), 0), _safe_int(row.get("lowest_quantity_after"), 0), _safe_int(row.get("current_quantity"), 0), _safe_int(row.get("threshold"), 0)]
        for row in data.get("problematic_products", [])
    ]
    for row in problematic_rows:
        ws_problematic.append(row)
    _style_ws(ws_problematic)

    ws_anomalies = wb.create_sheet("Аномалії")
    anomaly_rows = [["Товар", "SKU", "Аномалій", "Списань", "Коригувань", "Великих змін"]] + [
        [row.get("name") or "-", row.get("sku") or "-", _safe_int(row.get("anomaly_count"), 0), _safe_int(row.get("write_off_count"), 0), _safe_int(row.get("adjustment_count"), 0), _safe_int(row.get("large_delta_count"), 0)]
        for row in data.get("inventory_anomalies", [])
    ]
    for row in anomaly_rows:
        ws_anomalies.append(row)
    _style_ws(ws_anomalies)

    for ws in wb.worksheets:
        for column in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
            ws.column_dimensions[get_column_letter(column[0].column)].width = min(max(max_len + 2, 12), 42)
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)) and cell.column != 1:
                    cell.alignment = Alignment(horizontal="right", vertical="top")
        if ws.title in {"Замовлення", "Товари", "Склад"}:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=4 if ws.title != "Склад" else 5).alignment = Alignment(horizontal="center", vertical="top")

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()